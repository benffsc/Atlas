#!/usr/bin/env python3
"""
ingest_clinichq_historical.py

Ingests ClinicHQ historical XLSX reports into trapper.clinichq_hist_* tables.
Supports idempotent upserts via (source_file, source_row_hash) unique constraint.

Uses openpyxl for XLSX parsing (no pandas dependency).

Usage:
    # Ingest all three reports
    python3 ingest_clinichq_historical.py --all

    # Ingest specific report
    python3 ingest_clinichq_historical.py --appts
    python3 ingest_clinichq_historical.py --cats
    python3 ingest_clinichq_historical.py --owners

    # Dry run (no DB writes)
    python3 ingest_clinichq_historical.py --all --dry-run

    # Verbose mode
    python3 ingest_clinichq_historical.py --all --verbose

Requires:
    pip install openpyxl psycopg2-binary
"""

import argparse
import hashlib
import json
import re
import os
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Optional, Any, Iterator

# Try psycopg (v3)
try:
    import psycopg
except ImportError:
    print("ERROR: psycopg not found. Install with: pip install psycopg[binary]")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not found. Install with: pip install openpyxl")
    sys.exit(1)


# Default data paths (documented in NOW.md)
DATA_DIR = Path("data/clinichq/historical")
APPTS_FILE = DATA_DIR / "report_8af__appts.xlsx"
CATS_FILE = DATA_DIR / "report_82e__cats.xlsx"
OWNERS_FILE = DATA_DIR / "report_b38__owners.xlsx"

# Alternative path (legacy location)
ALT_DATA_DIR = Path("data/raw/clinichq/appointments/legacy_reports/8af_82e_b38")


def find_file(primary: Path, alt_dir: Path) -> Optional[Path]:
    """Find file in primary location or alternative directory."""
    if primary.exists():
        return primary
    alt_path = alt_dir / primary.name
    if alt_path.exists():
        return alt_path
    return None


def get_db_connection():
    """Get database connection from DATABASE_URL env var."""
    url = os.environ.get("DATABASE_URL")
    if not url:
        print("ERROR: DATABASE_URL environment variable not set")
        print("Run: set -a && source .env && set +a")
        sys.exit(1)
    return psycopg.connect(url)


def compute_row_hash(row_dict: dict) -> str:
    """Compute stable SHA256 hash of row data for idempotency."""
    serialized = json.dumps(row_dict, sort_keys=True, default=str)
    return hashlib.sha256(serialized.encode()).hexdigest()[:32]


def normalize_phone(raw_phone: Any) -> Optional[str]:
    """Normalize phone to E.164-ish format (digits only, +1 for US)."""
    if raw_phone is None:
        return None
    digits = ''.join(c for c in str(raw_phone) if c.isdigit())
    if not digits:
        return None
    if len(digits) == 10:
        return f"+1{digits}"
    elif len(digits) == 11 and digits[0] == '1':
        return f"+{digits}"
    elif len(digits) >= 7:
        return f"+{digits}"
    return None


def parse_date(val: Any) -> Optional[str]:
    """Parse date value to ISO string. Uses stdlib only (no dateutil)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    # Try common formats manually (stdlib only)
    str_val = str(val).strip()
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%m-%d-%Y',
                '%Y-%m-%d %H:%M:%S', '%m/%d/%Y %H:%M:%S',
                '%Y/%m/%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(str_val, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def parse_bool(val: Any) -> Optional[bool]:
    """Parse boolean-ish values."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    s = str(val).lower().strip()
    if s in ('true', 'yes', '1', 'x', 'checked'):
        return True
    if s in ('false', 'no', '0', '', 'unchecked'):
        return False
    return None


def parse_numeric(val: Any) -> Optional[float]:
    """Parse numeric values."""
    if val is None:
        return None
    try:
        return float(val)
    except:
        return None


def parse_int(val: Any) -> Optional[int]:
    """Parse integer values."""
    if val is None:
        return None
    try:
        return int(float(val))
    except:
        return None



def normalize_number(val):
    """Normalize ClinicHQ 'Number' fields. Keeps values like '14-1414' as text."""
    if val is None:
        return None
    txt = str(val).strip()
    if not txt or txt.lower() == "none":
        return None
    # If Excel gave a float that is actually an integer (e.g. 1234.0), strip .0
    if re.match(r"^\d+\.0$", txt):
        txt = txt[:-2]
    return txt

def iter_xlsx_rows(filepath: Path, verbose: bool = False) -> Iterator[dict]:
    """
    Iterate over XLSX rows as dictionaries.
    Uses openpyxl's read-only mode for memory efficiency.
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Get headers from first row
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(row)]
        break

    if verbose:
        print(f"  Headers: {headers[:5]}... ({len(headers)} total)")

    # Iterate data rows
    row_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val
        yield row_dict

        if verbose and row_count % 50000 == 0:
            print(f"    Progress: {row_count:,} rows read")

    wb.close()

    if verbose:
        print(f"  Total rows read: {row_count:,}")


def ingest_appts(dry_run: bool = False, verbose: bool = False) -> dict:
    """Ingest appointments from report_8af__appts.xlsx."""
    filepath = find_file(APPTS_FILE, ALT_DATA_DIR)
    if not filepath:
        print(f"ERROR: File not found: {APPTS_FILE}")
        print(f"       Also checked: {ALT_DATA_DIR / APPTS_FILE.name}")
        return {"error": "file_not_found"}

    print(f"Reading {filepath}...")
    source_file = filepath.name
    rows_to_insert = []

    for row_dict in iter_xlsx_rows(filepath, verbose):
        row_hash = compute_row_hash(row_dict)

        record = (
            parse_date(row_dict.get('Date')),
            normalize_number(row_dict.get('Number')),
            row_dict.get('Animal Name'),
            row_dict.get('Vet Name'),
            row_dict.get('Microchip Number'),
            row_dict.get('Internal Medical Notes'),
            row_dict.get('No Surgery Reason'),
            parse_bool(row_dict.get('Neuter')),
            parse_bool(row_dict.get('Spay')),
            parse_bool(row_dict.get('Cryptorchid')),
            parse_bool(row_dict.get('Pregnant')),
            parse_bool(row_dict.get('Pyometra')),
            parse_bool(row_dict.get('In Heat')),
            parse_bool(row_dict.get('URI')),
            parse_bool(row_dict.get('Fleas')),
            parse_bool(row_dict.get('Ticks')),
            parse_bool(row_dict.get('Ear mites')),
            parse_bool(row_dict.get('Tapeworms')),
            parse_bool(row_dict.get('Lactating')),
            parse_numeric(row_dict.get('Total Invoiced')),
            json.dumps(row_dict, default=str),
            source_file,
            row_hash
        )
        rows_to_insert.append(record)

    print(f"  Parsed {len(rows_to_insert):,} rows")

    if dry_run:
        print(f"  DRY RUN: Would insert {len(rows_to_insert):,} rows")
        return {"processed": len(rows_to_insert), "inserted": 0, "dry_run": True}

    print(f"  Inserting rows...")
    conn = get_db_connection()
    cur = conn.cursor()

    insert_sql = """
        INSERT INTO trapper.clinichq_hist_appts (
            appt_date, appt_number, animal_name, vet_name, microchip_number,
            internal_medical_notes, no_surgery_reason,
            neuter, spay, cryptorchid, pregnant, pyometra, in_heat,
            uri, fleas, ticks, ear_mites, tapeworms, lactating,
            total_invoiced, raw_row, source_file, source_row_hash
        ) VALUES %s
        ON CONFLICT (source_file, source_row_hash) DO NOTHING
    """

    # psycopg3: convert "VALUES %s" (psycopg2 execute_values style) into a normal single-row VALUES(...)
    if rows_to_insert:
        placeholders = ", ".join(["%s"] * len(rows_to_insert[0]))
        insert_sql = insert_sql.replace("VALUES %s", f"VALUES ({placeholders})")


    # Insert in batches
    batch_size = 1000
    total_inserted = 0
    for i in range(0, len(rows_to_insert), batch_size):
        batch = rows_to_insert[i:i+batch_size]
        cur.executemany(insert_sql, batch)
        total_inserted += cur.rowcount
        if verbose and (i + batch_size) % 10000 == 0:
            print(f"    Inserted: {min(i + batch_size, len(rows_to_insert)):,}/{len(rows_to_insert):,}")

    conn.commit()
    cur.close()
    conn.close()

    print(f"  Done: processed={len(rows_to_insert):,}, inserted={total_inserted:,}")
    return {"processed": len(rows_to_insert), "inserted": total_inserted}


def ingest_cats(dry_run: bool = False, verbose: bool = False) -> dict:
    """Ingest cats from report_82e__cats.xlsx."""
    filepath = find_file(CATS_FILE, ALT_DATA_DIR)
    if not filepath:
        print(f"ERROR: File not found: {CATS_FILE}")
        print(f"       Also checked: {ALT_DATA_DIR / CATS_FILE.name}")
        return {"error": "file_not_found"}

    print(f"Reading {filepath}...")
    source_file = filepath.name
    rows_to_insert = []
    bad_weight_count = 0
    bad_weight_examples = []

    for row_dict in iter_xlsx_rows(filepath, verbose):
        row_hash = compute_row_hash(row_dict)

        # Parse weight with guardrails for corrupt data
        weight_raw = row_dict.get('Weight')
        weight = parse_numeric(weight_raw)

        # Guardrail: cats should never be anywhere near 60+ lbs.
        # If outside range, treat as corrupt (often phone/reference numbers in Weight field).
        if weight is not None and (weight <= 0 or weight > 60):
            bad_weight_count += 1
            if len(bad_weight_examples) < 5:
                bad_weight_examples.append({
                    'raw': weight_raw,
                    'parsed': weight,
                    'animal': row_dict.get('Animal Name'),
                    'number': row_dict.get('Number')
                })
            weight = None

        record = (
            parse_date(row_dict.get('Date')),
            normalize_number(row_dict.get('Number')),
            row_dict.get('Animal Name'),
            row_dict.get('Microchip Number'),
            row_dict.get('Breed'),
            row_dict.get('Sex'),
            row_dict.get('Primary Color'),
            row_dict.get('Secondary Color'),
            row_dict.get('Spay Neuter Status'),
            weight,
            parse_int(row_dict.get('Age Months')),
            parse_int(row_dict.get('Age Years')),
            json.dumps(row_dict, default=str),
            source_file,
            row_hash
        )
        rows_to_insert.append(record)

    print(f"  Parsed {len(rows_to_insert):,} rows")
    if bad_weight_count:
        print(f"  NOTE: nulled {bad_weight_count:,} bad weights (kept in raw_row['Weight'])")
        if bad_weight_examples:
            print(f"  Examples of bad weights:")
            for ex in bad_weight_examples:
                print(f"    raw={ex['raw']} parsed={ex['parsed']} animal={ex['animal']} number={ex['number']}")

    if dry_run:
        print(f"  DRY RUN: Would insert {len(rows_to_insert):,} rows")
        return {"processed": len(rows_to_insert), "inserted": 0, "dry_run": True}

    print(f"  Inserting rows...")
    conn = get_db_connection()
    cur = conn.cursor()

    insert_sql = """
        INSERT INTO trapper.clinichq_hist_cats (
            appt_date, appt_number, animal_name, microchip_number,
            breed, sex, primary_color, secondary_color, spay_neuter_status,
            weight, age_months, age_years,
            raw_row, source_file, source_row_hash
        ) VALUES %s
        ON CONFLICT (source_file, source_row_hash) DO NOTHING
    """

    # psycopg3: convert "VALUES %s" (psycopg2 execute_values style) into a normal single-row VALUES(...)
    if rows_to_insert:
        placeholders = ", ".join(["%s"] * len(rows_to_insert[0]))
        insert_sql = insert_sql.replace("VALUES %s", f"VALUES ({placeholders})")


    batch_size = 1000
    total_inserted = 0
    for i in range(0, len(rows_to_insert), batch_size):
        batch = rows_to_insert[i:i+batch_size]
        cur.executemany(insert_sql, batch)
        total_inserted += cur.rowcount

    conn.commit()
    cur.close()
    conn.close()

    print(f"  Done: processed={len(rows_to_insert):,}, inserted={total_inserted:,}")
    return {"processed": len(rows_to_insert), "inserted": total_inserted}


def ingest_owners(dry_run: bool = False, verbose: bool = False) -> dict:
    """Ingest owners from report_b38__owners.xlsx."""
    filepath = find_file(OWNERS_FILE, ALT_DATA_DIR)
    if not filepath:
        print(f"ERROR: File not found: {OWNERS_FILE}")
        print(f"       Also checked: {ALT_DATA_DIR / OWNERS_FILE.name}")
        return {"error": "file_not_found"}

    print(f"Reading {filepath}...")
    source_file = filepath.name
    rows_to_insert = []

    for row_dict in iter_xlsx_rows(filepath, verbose):
        row_hash = compute_row_hash(row_dict)

        # Normalize phone (prefer cell, fall back to landline)
        phone_norm = normalize_phone(row_dict.get('Owner Cell Phone')) or \
                     normalize_phone(row_dict.get('Owner Phone'))

        record = (
            parse_date(row_dict.get('Date')),
            normalize_number(row_dict.get('Number')),
            row_dict.get('Animal Name'),
            row_dict.get('Microchip Number'),
            row_dict.get('Ownership'),
            row_dict.get('Owner First Name'),
            row_dict.get('Owner Last Name'),
            row_dict.get('Owner Address'),
            row_dict.get('Owner Cell Phone'),
            row_dict.get('Owner Phone'),
            row_dict.get('Owner Email'),
            row_dict.get('ClientType'),
            phone_norm,
            json.dumps(row_dict, default=str),
            source_file,
            row_hash
        )
        rows_to_insert.append(record)

    print(f"  Parsed {len(rows_to_insert):,} rows")

    if dry_run:
        print(f"  DRY RUN: Would insert {len(rows_to_insert):,} rows")
        return {"processed": len(rows_to_insert), "inserted": 0, "dry_run": True}

    print(f"  Inserting rows...")
    conn = get_db_connection()
    cur = conn.cursor()

    insert_sql = """
        INSERT INTO trapper.clinichq_hist_owners (
            appt_date, appt_number, animal_name, microchip_number,
            ownership, owner_first_name, owner_last_name, owner_address,
            owner_cell_phone, owner_phone, owner_email, client_type,
            phone_normalized,
            raw_row, source_file, source_row_hash
        ) VALUES %s
        ON CONFLICT (source_file, source_row_hash) DO NOTHING
    """

    # psycopg3: convert "VALUES %s" (psycopg2 execute_values style) into a normal single-row VALUES(...)
    if rows_to_insert:
        placeholders = ", ".join(["%s"] * len(rows_to_insert[0]))
        insert_sql = insert_sql.replace("VALUES %s", f"VALUES ({placeholders})")


    batch_size = 1000
    total_inserted = 0
    for i in range(0, len(rows_to_insert), batch_size):
        batch = rows_to_insert[i:i+batch_size]
        cur.executemany(insert_sql, batch)
        total_inserted += cur.rowcount

    conn.commit()
    cur.close()
    conn.close()

    print(f"  Done: processed={len(rows_to_insert):,}, inserted={total_inserted:,}")
    return {"processed": len(rows_to_insert), "inserted": total_inserted}


def main():
    parser = argparse.ArgumentParser(
        description="Ingest ClinicHQ historical XLSX reports (no pandas)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Files expected in: data/clinichq/historical/
  - report_8af__appts.xlsx  (272K appointments)
  - report_82e__cats.xlsx   (37K cat records)
  - report_b38__owners.xlsx (37K owner records)

Also checks: data/raw/clinichq/appointments/legacy_reports/8af_82e_b38/
        """
    )
    parser.add_argument("--all", action="store_true", help="Ingest all three reports")
    parser.add_argument("--appts", action="store_true", help="Ingest appointments (report_8af)")
    parser.add_argument("--cats", action="store_true", help="Ingest cats (report_82e)")
    parser.add_argument("--owners", action="store_true", help="Ingest owners (report_b38)")
    parser.add_argument("--dry-run", action="store_true", help="Parse files but don't write to DB")
    parser.add_argument("--verbose", action="store_true", help="Show progress during parsing")

    args = parser.parse_args()

    if not (args.all or args.appts or args.cats or args.owners):
        parser.print_help()
        print("\nNo action specified. Use --all or --appts/--cats/--owners")
        sys.exit(1)

    results = {}

    if args.all or args.appts:
        print("\n=== APPOINTMENTS (report_8af) ===")
        results['appts'] = ingest_appts(dry_run=args.dry_run, verbose=args.verbose)

    if args.all or args.cats:
        print("\n=== CATS (report_82e) ===")
        results['cats'] = ingest_cats(dry_run=args.dry_run, verbose=args.verbose)

    if args.all or args.owners:
        print("\n=== OWNERS (report_b38) ===")
        results['owners'] = ingest_owners(dry_run=args.dry_run, verbose=args.verbose)

    print("\n=== SUMMARY ===")
    for name, result in results.items():
        if 'error' in result:
            print(f"  {name}: ERROR - {result['error']}")
        elif result.get('dry_run'):
            print(f"  {name}: DRY RUN - would process {result['processed']:,} rows")
        else:
            print(f"  {name}: processed={result['processed']:,}, inserted={result['inserted']:,}")

    # Exit with error if any failed
    if any('error' in r for r in results.values()):
        sys.exit(1)


if __name__ == "__main__":
    main()
